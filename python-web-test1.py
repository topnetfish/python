#!/usr/bin/python
# -*- coding: UTF-8 -*-
__author__ = 'z00301276'
import requests
import openpyxl
#import json
def Geturldata():
    data = {'actionFlag':'loginAuthenticate', \
            'redirect':'http://w3.huawei.com/info/cn/browseBPA.do?id=19981', \
            'uid':'z00301276', \
            'password':'zhou@2019'
            }
    session = requests.session()
    r = session.post('https://login.huawei.com/login/login.do', data = data)
    if r.status_code != 200:
        print("connection error, the errorcode is:%s" %r.status_code)
        return None

    paras = {"method":"listPageDocPublishByCondition", \
             "dataScrope":"4", \
             "id": "19981", \
             "docstatus":"-1", \
             "pagesize":"20", \
             "curpage":"1", \
             "key":"5G"}
    response = session.get('http://w3.huawei.com/info/cn/viewBPA.do', params = paras)
    #print(response.json())
    if response.status_code == 200:
        return response
    else:
        print("connection error, the errorcode is:%s" %response.status_code)
        return None

def WriteData(response):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = "标题"
    ws['B1'] = "链接"
    ws['C1'] = "发布部门"
    ws['D1'] = "发布时间"
    i = 2
    basic_url = "http://w3.huawei.com/info/cn/doc/viewDoc.do?did="
    for content_dict in response.json()["ajaxData"]["items"]:
        ws.cell(row = i, column = 1, value = content_dict["name"])
        ws.cell(row = i, column = 2, value = basic_url + str(content_dict["id"]))
        ws.cell(row = i, column = 3, value = content_dict["publishDept"])
        ws.cell(row = i, column = 4, value = content_dict["publishedDate"])
        i += 1
    #保存在程序所在路径下，也可以指定保存到别的路径下
    wb.save("5G喜报_z00301276.xlsx")
    print("5G save ok ",i)

response = Geturldata()
if None != response:
    WriteData(response)